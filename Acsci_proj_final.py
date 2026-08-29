import numpy as np
import pandas as pd
import re as re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from scipy.optimize import minimize
from scipy.interpolate import interp1d
from openpyxl.styles import borders
from openpyxl.chart import BarChart, Reference

def clean_currency_text(value):
    if pd.isna(value):
        return np.nan
    if isinstance(value, str):
        cleaned = re.sub(r"[^\d.]", "", value)
        return float(cleaned if cleaned else np.nan)
    return float(value)


def load_policy_book(filepath):
    df = pd.read_excel(filepath)

    required_cols = {"policy_id", "issue_age", "face_amount", "annual_premium"}
    missing_cols = required_cols - set(df.columns)
    assert not missing_cols, f"Missing required columns: {missing_cols}"

    df["policy_id"] = df["policy_id"].astype(str).str.strip().str.upper()
    df["face_amount"] = df["face_amount"].apply(clean_currency_text)
    df["annual_premium"] = df["annual_premium"].apply(clean_currency_text)
    df["issue_age"] = pd.to_numeric(df["issue_age"], errors="coerce").astype("Int64")

    bad_rows = df[df[list(required_cols)].isnull().any(axis=1)]
    if len(bad_rows) > 0:
        print(f"Dropping {len(bad_rows)} row(s) with missing required data after cleaning:")
        print(bad_rows)
    df = df.dropna(subset=list(required_cols))

    return df.reset_index(drop=True)


def get_qx_for_policy(mortality_rates, issue_age, n_years=None):
    if n_years is None:
        n_years = len(mortality_rates) - issue_age
    qx = mortality_rates[issue_age: issue_age + n_years]
    assert len(qx) == n_years, f"Only got {len(qx)} mortality rates for issue age {issue_age}, expected {n_years}"
    return qx


def implied_net_level_premium(mortality_rates, extended_lapse, issue_age, face_amount, interest_rate):
    qx = get_qx_for_policy(mortality_rates, issue_age)
    n_years = len(qx)
    lapse = extended_lapse[:n_years]
    combined_px = [(1 - q) * (1 - l) for q, l in zip(qx, lapse)]
    tpx_start = [1.0] + list(np.cumprod(combined_px)[:-1])
    v = [1 / (1 + interest_rate) ** t for t in range(n_years)]

    apv_benefits = sum(tpx * qx_ * face_amount * v_ for tpx, qx_, v_ in zip(tpx_start, qx, v))
    apv_premium_annuity = sum(tpx * v_ for tpx, v_ in zip(tpx_start, v))

    return apv_benefits / apv_premium_annuity

def calibrate_book_premiums(policy_book, mortality_rates, extended_lapse, interest_rate, loading= 1.0):
    calibrated = policy_book.copy()
    calibrated["annual_premium"] = [
        round(implied_net_level_premium(mortality_rates, extended_lapse, age, face, interest_rate) * loading, 2)
        for age, face in zip(calibrated["issue_age"], calibrated["face_amount"])
    ]
    return calibrated

def build_extended_lapse_table(base_lapse, total_years, ultimate_rate, decay_speed=0.15):
    extended = list(base_lapse)
    last_known = extended[-1]

    for t in range(len(extended), total_years):
        years_past_known = t - len(base_lapse)
        rate = ultimate_rate + (last_known - ultimate_rate) * np.exp(-decay_speed * years_past_known)
        extended.append(rate)
    return extended


def randomize_mortality(qx_policy, shock_std_mort):
    return [qx * np.random.lognormal(mean=0, sigma=shock_std_mort) for qx in qx_policy]


def randomize_lapse(base_lapse, shock_std_lapse):
    return [max(0, min(l + np.random.normal(0, shock_std_lapse), 1)) for l in base_lapse]


def randomize_lapse_dynamic(base_lapse, shock_std_lapse, current_rate, long_run_mean, sensitivity=2.0, idiosyncratic_std = 0.002):
    rate_deviation = current_rate - long_run_mean
    dynamic_shift = sensitivity * rate_deviation
    policy_noise = np.random.normal(0, idiosyncratic_std)
    return [max(0, min(l + dynamic_shift + policy_noise + np.random.normal(0, shock_std_lapse), 1)) for l in base_lapse]

def benefit_only_pv(qx_policy, base_lapse, shock_std_mort, policy_face, discount_factors, radix=100000):
    sim_mort = randomize_mortality(qx_policy, shock_std_mort)
    combined_px = [(1 - qx) * (1 - lapse) for qx, lapse in zip(sim_mort, base_lapse)]
    survival_prob_local = np.cumprod(combined_px)
    lx_start = [radix] + [radix * s for s in survival_prob_local[:-1]]
    tpx_start = [l / radix for l in lx_start]
    expected_benefit = [tpx * qx * policy_face for tpx, qx in zip(tpx_start, sim_mort)]
    return sum(b * d for b, d in zip(expected_benefit, discount_factors))

def simulate_rate_path_vasicek(start_rate, long_run_mean, speed, annual_vol, n_years):
    rates = [start_rate]
    for _ in range(n_years - 1):
        prev = rates[-1]
        shock = np.random.normal(0, annual_vol)
        rates.append(prev + speed * (long_run_mean - prev) + shock)
    return rates


def simulate_n_paths(n_simulation, start_rate, long_run_mean, speed, annual_vol, n_years):
    all_paths = []
    for _ in range(n_simulation):
        all_paths.append(simulate_rate_path_vasicek(start_rate, long_run_mean, speed, annual_vol, n_years))
    return np.array(all_paths)


def reserve_for_one_path(reserve_factor_cv, reserve_factor_prem, discount_factors):
    pv_benefits = sum(b * d for b, d in zip(reserve_factor_cv, discount_factors))
    pv_premiums = sum(p * d for p, d in zip(reserve_factor_prem, discount_factors))
    return pv_benefits - pv_premiums


def simulate_one_reserve(qx_policy, base_lapse, shock_std_mort, shock_std_lapse, policy_face, annual_premium, discount_factors, current_rate, long_run_mean, radix=100000, lapse_sensitivity=2.0):
    sim_mort = randomize_mortality(qx_policy, shock_std_mort)
    sim_lapse = randomize_lapse_dynamic(base_lapse, shock_std_lapse, current_rate, long_run_mean, sensitivity=lapse_sensitivity)
    combined_px = [(1 - qx) * (1 - lapse) for qx, lapse in zip(sim_mort, sim_lapse)]
    survival_prob_local = np.cumprod(combined_px)
    lx_start = [radix] + [radix * s for s in survival_prob_local[:-1]]
    tpx_start = [l / radix for l in lx_start]
    expected_benefit = [tpx * qx * policy_face for tpx, qx in zip(tpx_start, sim_mort)]
    expected_premiums = [tpx * annual_premium for tpx in tpx_start]
    return reserve_for_one_path(expected_benefit, expected_premiums, discount_factors)


def simulate_one_book_reserve(policy_book, mortality_rates, base_lapse, shock_std_mort, shock_std_lapse, discount_factors, n_years, current_rate, long_run_mean, radix=100000):
    total_reserve = 0
    for policy in policy_book.to_dict('records'):
        qx_policy = get_qx_for_policy(mortality_rates, policy["issue_age"], n_years)
        reserve = simulate_one_reserve(qx_policy, base_lapse, shock_std_mort, shock_std_lapse, policy["face_amount"], policy["annual_premium"], discount_factors, current_rate, long_run_mean, radix)
        total_reserve += max(0, reserve)
    return total_reserve


def run_full_book_simulation_lifetime(n_sims, start_rate, long_run_mean, speed, annual_vol, policy_book, mortality_rates, extended, shock_std_mort, shock_std_lapse, warn_threshold=1000):
    if n_sims > warn_threshold or len(policy_book) > 500:
        print(f"Warning: {n_sims} sims x {len(policy_book)} policies may take several minutes. Consider a smaller test run first.")

    max_horizon = len(mortality_rates) - policy_book["issue_age"].min()
    paths = simulate_n_paths(n_sims, start_rate, long_run_mean, speed, annual_vol, max_horizon)
    discount_paths = np.array([np.cumprod(1/(1+i)) for i in paths])

    book_reserves = []
    for path, d in zip(paths, discount_paths):
        current_rate = float(np.mean(path))
        total = 0
        for policy in policy_book.to_dict('records'):
            policy_n_years = len(mortality_rates) - policy["issue_age"]
            qx_policy = get_qx_for_policy(mortality_rates, policy["issue_age"], policy_n_years)
            policy_discount = d[:policy_n_years]
            lapse_policy = (extended + [extended[-1]] * (policy_n_years - len(extended)))[:policy_n_years]
            reserve = simulate_one_reserve(qx_policy, lapse_policy, shock_std_mort, shock_std_lapse, policy["face_amount"], policy["annual_premium"], policy_discount, current_rate, long_run_mean)
            total += max(0, reserve)
        book_reserves.append(total)

    return np.array(book_reserves)


def run_full_simulation(n_sims, start_rate, long_run_mean, speed, annual_vol, n_years, qx_policy, base_lapse, shock_std_mort, shock_std_lapse, policy_face, annual_premium):
    paths = simulate_n_paths(n_sims, start_rate, long_run_mean, speed, annual_vol, n_years)
    discount_paths = np.array([np.cumprod(1/(1+i)) for i in paths])
    return np.array([simulate_one_reserve(qx_policy, base_lapse, shock_std_mort, shock_std_lapse, policy_face, annual_premium, d, float(np.mean(p)), long_run_mean) for p, d in zip(paths, discount_paths)])


def run_full_book_simulation(n_sims, start_rate, long_run_mean, speed, annual_vol, n_years, policy_book, mortality_rates, base_lapse, shock_std_mort, shock_std_lapse):
    paths = simulate_n_paths(n_sims, start_rate, long_run_mean, speed, annual_vol, n_years)
    discount_paths = np.array([np.cumprod(1/(1+i)) for i in paths])
    return np.array([simulate_one_book_reserve(policy_book, mortality_rates, base_lapse, shock_std_mort, shock_std_lapse, d, n_years, float(np.mean(p)), long_run_mean) for p, d in zip(paths, discount_paths)])


def export_results_to_excel(filepath, policy_book, book_reserves, assumptions_dict):
    wb = Workbook()

    CUR = '$#,##0.00;($#,##0.00)'
    PCT = '0.00%'
    Header_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    Header_fill = PatternFill("solid", start_color="1F4E78")
    Label = Font(name="Arial", bold=True, size=10)
    chart = BarChart()
    chart.title = "Book Reserve Distribution"
    chart.y_axis.title = "Frequency"
    chart.x_axis.title = "Reserve ($)"
    pct_keys = {"start_rate", "long_run_mean", "speed", "annual_vol", "shock_std_mort", "shock_std_lapse"}

    ws = wb.active
    ws.title = "Assumptions"
    ws["A1"] = "Model Assumptions"
    ws["A1"].font = Header_font
    ws["A1"].fill = Header_fill
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 18

    r = 3
    for label, val in assumptions_dict.items():
        ws.cell(row=r, column=1, value=label).font = Label
        value_cell = ws.cell(row=r, column=2, value=val)
        if label in pct_keys: value_cell.number_format = PCT
        r += 1

    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells("A1:B1")
    ws2["A1"] = "Simulation Results Summary"
    ws2["A1"].font = Header_font
    ws2["A1"].fill = Header_fill
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 20

    VaR_95 = np.percentile(book_reserves, 95)
    tail = book_reserves[book_reserves >= VaR_95]
    CVaR_95 = tail.mean()

    summary_items = [
        ("Number of policies", len(policy_book)),
        ("Number of simulations", len(book_reserves)),
        ("Mean book reserve", round(book_reserves.mean(), 2)),
        ("Std dev of book reserve", round(book_reserves.std(), 2)),
        ("Min book reserve", round(book_reserves.min(), 2)),
        ("Max book reserve", round(book_reserves.max(), 2)),
        ("95% VaR", round(VaR_95, 2)),
        ("95% CVaR", round(CVaR_95, 2)),
    ]

    r = 3
    for label, val in summary_items:
        ws2.cell(row=r, column=1, value=label).font = Label
        value_cell = ws2.cell(row=r, column=2, value=val)
        if "reserve" in label.lower() or "var" in label.lower() or "cvar" in label.lower():
            value_cell.number_format = CUR
        r += 1

    ws3 = wb.create_sheet("Raw Simulation Results")
    ws3.freeze_panes = "A2"
    ws3["A1"] = "Simulation #"
    ws3["B1"] = "Book Reserve"
    ws3["A1"].font = Label
    ws3["B1"].font = Label
    for i, val in enumerate(book_reserves, start=1):
        ws3.cell(row=i+1, column=1, value=i)
        ws3.cell(row=i+1, column=2, value=round(float(val), 2))

    ws4 = wb.create_sheet("Policy_Book")
    cols = list(policy_book.columns)
    for c, col_name in enumerate(cols, start=1):
        ws4.cell(row=1, column=c, value=col_name).font = Label
    for r, row in enumerate(policy_book.itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            ws4.cell(row=r, column=c, value=val)

    bins = np.histogram(book_reserves, bins=15)
    ws5 = wb.create_sheet("Chart Data")
    for i, (count, edge) in enumerate(zip(bins[0], bins[1]), start=1):
        ws5.cell(row=i, column=1, value=f"${edge:,.0f}")
        ws5.cell(row=i, column=2, value=int(count))

    data = Reference(ws5, min_col=2, min_row=1, max_row=len(bins[0]))
    cats = Reference(ws5, min_col=1, min_row=1, max_row=len(bins[0]))
    chart.add_data(data)
    chart.set_categories(cats)
    ws2.add_chart(chart, "D3")
    wb.save(filepath)
    return filepath
